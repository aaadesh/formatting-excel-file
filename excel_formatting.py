from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Alignment, Font

#H:\Git\Excel Formatting\patent data.xlsx
source = input("Enter the path of file:")

wb = load_workbook(source)
ws = wb.active

#print(ws)

ws.insert_rows(1)
ws.insert_cols(1)

ws.row_dimensions[1].height = 5
ws.column_dimensions['A'].width = 1


maxr = ws.max_row 
maxc = ws.max_column 

minr = ws.min_row 
minc = ws.min_column 
#print(minr, minc, maxr, maxc)


# setting fonts and fills
header_font = Font(name='Arial',
                 size=10,
                 bold=True,
                 italic=False,
                 vertAlign=None,
                 color='00FFFFFF')
header_fill = PatternFill(start_color='00000080',
                   end_color='FFFF0000',
                   fill_type='solid')

text_font = Font(name='Arial',
                 size=10,
                 bold=False,
                 italic=False,
                 vertAlign=None,
                 color='00000000')
add_border = Border(left=Side(border_style=None,
                           color='FF000000'),
                 right=Side(border_style=None,
                            color='FF000000'),
                 top=Side(border_style='thin',
                          color='00C0C0C0'),
                 bottom=Side(border_style='thin',
                             color='00C0C0C0'),
                 diagonal=Side(border_style=None,
                               color='FF000000'),
                 diagonal_direction=0,
                 outline=Side(border_style=None,
                              color='FF000000'),
                 vertical=Side(border_style=None,
                               color='FF000000'),
                 horizontal=Side(border_style=None,
                                color='FF000000')
                )


# set font, fill, and alignment of header
for i in range(minc, maxc +1):
    #print(ws.cell(row = 1, column = i).value)
    ws.cell(row = minr +1, column = i).font = header_font
    ws.cell(row = minr +1, column = i).fill = header_fill
    ws.cell(row = minr +1, column = i).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

blankr = []
blankc = []

# set font, and alignement of other text
for j in range(minc, maxc +1):
    for k in range(minr +2, maxr +1):
        #print(ws.cell(row = k, column = j).value)
        ws.cell(row = k, column = j).font = text_font
        ws.cell(row = k, column = j).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        ws.cell(row = k, column = j).border = add_border
        if ws.cell(row = k, column = j).value == None:
            print(k, j)
            blankr.append(k)
            blankc.append(j)
            

# to replace blank columns with hyphen
for n in range(len(blankr)):
    ws.cell(row = blankr[n], column = blankc[n]).value = "-"


# set the height of the row
for l in range(minr +1, maxr +1):
    ws.row_dimensions[l].height = 25
  
# set the width of the column
for m in range(minc +1, maxc +1):
    ws.column_dimensions[str(chr(64 + m))].width = 15


# to remove gridlines
ws.sheet_view.showGridLines = False

wb.save("formatted file.xlsx")
